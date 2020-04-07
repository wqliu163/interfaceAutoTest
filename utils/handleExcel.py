from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from utils.log.excuteWriteLog import *
import os
import traceback


class handleExcel(object):
    colourDict={'red': 'FF0000', 'green': '00FF00','black':'000000','gray':'CCCCCC','yellow':'FFFF00','white':'FFFFFF'}
    @classmethod
    def createExcelTable(cls,path,tableName):
        """
        :param path:        (str)     文件所在路径
        :param tableName:   (str)    文件名称，包含后缀名，如2005.xlsx
        :return:            (object)  excel 对象
        """
        workBook=Workbook()
        excelPath=path.strip("\\")+"\\"+tableName
        if not os.path.exists(excelPath):
            workBook.save(excelPath)
            return excelPath
        else:
            print("the path exist file of the name !")
            workBook=load_workbook(excelPath)
            return workBook

    @classmethod
    def loadExcelTable(cls,path):
        """
        :param path:        (str)       文件所在路径
        :return:            (object)     excel 对象
        """
        if os.path.exists(path) and os.path.isfile(path):
            if os.path.splitext(".xlsx") or os.path.splitext(".xls"):
                workBook=load_workbook(path)
                return workBook
            else:
                print("the path not exist !")
                return None
        else:
            print("the path not exist !")
            return None

    @classmethod
    def createExcelSheet(cls,path,sheetName,NO):
        """
        :param path:        (str)       文件所在路径
        :param sheetName:   (str)       工作表名
        :param NO:          (int)       插入位置，如0，最前面
        :return:            (object)     excel 对象
        """
        workBook=handleExcel.loadExcelTable(path)
        if sheetName not in workBook.sheetnames:
            workBook.create_sheet(sheetName,NO)
        else:
            print("the table exist sheet of the name !")
        #workBook.save(path)
        return workBook

    @classmethod
    def deleteExcelSheet(cls,path,sheetName):
        """
        :param path:         (str)      文件所在路径
        :param sheetName:    (str)      工作表名
        :return:            (object)    excel 对象
        """
        workBook=handleExcel.loadExcelTable(path)
        del workBook[sheetName]
        #workBook.save(path)
        return workBook

    @classmethod
    def getSheetObjectByName(cls,path,sheetName):
        """
        :param path:         (str)       文件所在路径
        :param sheetName:    (str)       工作表名
        :return:             (object)    对象 workBook,workSheet
        """
        workBook=handleExcel.loadExcelTable(path)
        if sheetName in workBook.sheetnames:
            workSheet = workBook[sheetName]
            return workBook,workSheet
        else:
            print("the table not exist sheet of the name !")
            return None

    @classmethod
    def getSheetObjectByIndex(cls,path,sheetIndex):
        """
        :param path:         (str)      文件所在路径
        :param sheetIndex:   (int)      工作表位置号
        :return:            (object)    对象 workBook,workSheet
        """
        workBook=handleExcel.loadExcelTable(path)
        if isinstance(sheetIndex,int)  and sheetIndex>=0:
            if len(workBook.sheetnames)-1 >=sheetIndex:
                workSheet=workBook[workBook.sheetnames[sheetIndex]]
                return workBook,workSheet
            else:
                print("out of index range")
                return None
        else:
            print("the sheetIndex parameter error !")
            return None

    @classmethod
    def getMinColumnNO(cls,path,sheetName):
        """
        :param path:        (str)      文件所属地址
        :param sheetName:   (str)       工作表名
        :return:            (int)      有效最小列号
        """
        workBook,workSheet=handleExcel.getSheetObjectByName(path,sheetName)
        print(workSheet.min_column)
        return workSheet.min_column

    @classmethod
    def getMaxColumnNO(cls, path, sheetName):
        """
        :param path:        (str)      文件所属地址
        :param sheetName:   (str)      工作表名
        :return:            (int)      有效最大列号
        """
        workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
        print(workSheet.max_column)
        return workSheet.max_column

    @classmethod
    def getMinRowNO(cls, path, sheetName):
        """
        :param path:        (str)      文件所属地址
        :param sheetName:   (str)       工作表名
        :return:            (int)      有效最小行号
        """
        workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
        print(workSheet.min_row)
        return workSheet.min_row

    @classmethod
    def getMaxRowNO(cls, path, sheetName):
        """
        :param path:       (str)   文件所属地址
        :param sheetName:  (str)   工作表名
        :return:           (int)    有效最大行号
        """
        workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
        print(workSheet.max_row)
        return workSheet.max_row

    @classmethod
    def getRowsObject(cls,path,sheetName,rowNo):
        """
        :param path:         (str)   文件所属地址
        :param sheetName:    (str)   工作表名
        :param rowNo:        (int)   行号
        :return:            (object)    行对象
        """
        if isinstance(rowNo,int) and rowNo>=0:
            workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
            rowObject=workSheet[rowNo]
            print(rowObject)
            return rowObject
        else:
            print("parameter error !,rowNo type must be int! ")
            return None

    @classmethod
    def getColumnsObject(cls,path,sheetName,columnNo):
        """
        :param path:        (str)           文件所属地址
        :param sheetName:   (str)           工作表名
        :param columnNo:   (int or str)     行号 1-26||a-z||A-Z
        :return:           (object)         行对象
        """
        workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
        if isinstance(columnNo,str):
            try:
                columnObject=workSheet[columnNo]
                print(columnObject)
                return columnObject
            except Exception as error:
                excutelog("error", traceback.format_exc())
                print(error)
        elif isinstance(columnNo,int) and columnNo>0 and columnNo<=26:
            columnObject = workSheet[chr(columnNo+64)]
            print(columnObject)
            return columnObject
        elif isinstance(columnNo,int) and columnNo>26:
            columnNoNew=chr(columnNo//26+64)+chr(columnNo%26+64)
            columnObject=workSheet[columnNoNew]
            print(columnObject)
            return columnObject
        else:
            print("parameter error !,columnNo type must be str 'a-z || A-Z' or int >0 ! ")
            return None

    @classmethod
    def getValueOfCell(cls,path,sheetName,coordinate=None,rowNo=None,columnNo=None):
        """
        :param path:        (str)   文件所属路径
        :param sheetName:   (str)   工作表名
        :param coordinate:  (str)   单元格坐标 如A1,B5
        :param rowNo:       (int)   行号
        :param columnNo:    (int)    列号
        :return:
        """
        workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
        if coordinate !=None and isinstance(coordinate,str):
            try:
                cellValue=workSheet[coordinate].value
                print(cellValue)
                return cellValue
            except Exception as error:
                excutelog("error",traceback.format_exc())
                print(error)
        elif coordinate ==None and isinstance(rowNo,int) and isinstance(columnNo,int):
            try:
                cellValue = workSheet.cell(row=rowNo,column=columnNo).value
                print(cellValue)
                return cellValue
            except Exception as error:
                excutelog("error", traceback.format_exc())
                print(error)
        else:
            print("parameter error,coordinate: (str)单元格坐标 如A1,B5;rowNo:(int)行号;columnNo:(int)列号")
            return None
    @classmethod
    def writeValueInCell(cls,path,sheetName,values,coordinate=None,rowNo=None,columnNo=None,colour="black"):
        """
        :param path:        (str)       文件所属路径
        :param sheetName:   (str)   工作表名
        :param values:              写入的内容
        :param coordinate:  (str)   单元格坐标 如A1,B5
        :param rowNo:       (int)   行号
        :param columnNo:    (int)    列号
        :param colour:      (str)    颜色，colourDict中的值，如"red","black"
        :return:
        """
        workBook,workSheet = handleExcel.getSheetObjectByName(path, sheetName)
        try:
            if coordinate!=None:
                workSheet[coordinate].value=values
                workSheet[coordinate].font=Font(color=handleExcel.colourDict[colour])
                workBook.save(path)
                return workBook,workSheet
            elif coordinate==None and isinstance(rowNo,int) and isinstance(columnNo,int) and rowNo>0 and columnNo>0:
                workSheet.cell(row=rowNo,column=columnNo).value=values
                workSheet.cell(row=rowNo, column=columnNo).font=Font(color=handleExcel.colourDict[colour])
                workBook.save(path)
                return workBook, workSheet
            else:
                if not isinstance(rowNo,int) or not isinstance(columnNo,int):
                    print("parameter columnNo or rowNo error !  columnNo or rowNo must be int type")
                else:
                    print("parameter coordinate error !  coordinate must be str type")
        except Exception as error:
            print(error)

if __name__=="__main__":
    #handleExcel.createExcelTable(r"C:\Users\liuwq\Desktop\pp","pp.xlsx")
    #handleExcel.loadExcelTable(r"K:\eDiary\learn_Eenlish.xlsx")
    #handleExcel.createExcelSheet(r"K:\eDiary\learn_Eenlish.xlsx","wqliu1",0)
    #handleExcel.deleteExcelSheet(r"K:\eDiary\learn_Eenlish.xlsx","wqliu")
    #handleExcel.getSheetObjectByName(r"K:\eDiary\learn_Eenlish.xlsx","wqliu")
    #handleExcel.getSheetObjectByIndex(r"K:\eDiary\learn_Eenlish.xlsx",0)
    #handleExcel.getMinColumnNO(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1")
    #handleExcel.getMaxColumnNO(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1")
    #handleExcel.getMinRowNO(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1")
    #handleExcel.getMaxRowNO(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1")
    #handleExcel.getRowsObject(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1",27)
    #handleExcel.getColumnsObject(getRootPath()+"\\data\\case.xlsx","API","G")
    #handleExcel.getValueOfCell(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1",coordinate="A1")
    handleExcel.getValueOfCell(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1",rowNo=2,columnNo="A")
    # workbook,sheetbook=handleExcel.writeValueInCell(r"K:\eDiary\learn_Eenlish.xlsx","Sheet1","good",rowNo=2,columnNo=14)
    # workbook.save(r"K:\eDiary\learn_Eenlish.xlsx")

